import openpyxl
import math
import matplotlib.pyplot as plt
from scipy.stats import norm
import statistics
import numpy as np

if __name__ == '__main__':

    # odczytanie danych
    dataframe = openpyxl.load_workbook("dane.xlsx")
    dataframe1 = dataframe.active

    # tablica na dane
    data = []

    # liczba zmiennych
    n = dataframe1.max_row-1

    # przypisanie danych do tablicy
    for row in range(1, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            data.append(col[row].value)

    # liczenie sredniej
    sum = 0
    for x in range(0, n):
        sum += data[x]

    avg = sum/n

    # odchylenie standardowe
    sum = 0
    for i in range(0, n):
        sum += pow((data[i] - avg), 2)
    error = math.sqrt(sum/n)

    # wzor gaussa
    xdata = []
    x = 1
    while x < 5.99:
        temp = (1 / (error * math.sqrt(2*math.pi))) * math.exp(((-1)*pow((x - avg), 2))/(2 * pow(error, 2)))
        xdata.append(temp)
        x += 0.01

    xdata2 = []
    x = 1
    while x < 7:
        temp = (1 / (error * math.sqrt(2 * math.pi))) * math.exp(((-1) * pow((x - avg), 2)) / (2 * pow(error, 2)))
        xdata2.append(temp)
        x += 1

    # wypisanie danych
    print(f"dane: {data}")
    print(f"liczba zmiennych: {n}")
    print(f"srednia: {avg}")
    print(f"error: {error}")

    x_axis = np.arange(1, 6, 0.01)

    # rozklad gaussa
    plt.plot(x_axis, norm.pdf(x_axis, statistics.mean(x_axis), statistics.stdev(x_axis)))

    # punkty nałożone na wykres
    plt.scatter(1, xdata2[0], color='red')
    plt.scatter(2, xdata2[1], color='red')
    plt.scatter(3, xdata2[2], color='red')
    plt.scatter(4, xdata2[3], color='red')
    plt.scatter(5, xdata2[4], color='red')
    plt.scatter(6, xdata2[5], color='red')
    plt.show()